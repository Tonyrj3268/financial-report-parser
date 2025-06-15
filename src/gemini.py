import base64
import json
import os
import tempfile
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
from pathlib import Path

import fitz
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from models.cash_equivalents import CashAndEquivalents, cash_equivalents_prompt
from models.prepayments import PrePayments, prepayments_prompt
from models.receivables_related_parties import (
    ReceivablesRelatedParties,
    receivables_related_parties_prompt,
)
from models.total_liabilities import TotalLiabilities, total_liabilities_prompt
from models.corporate_bond_payable import (
    CorporateBondPayable,
    corporate_bond_payable_prompt,
)
from models.property_plant_equipment import (
    PropertyPlantEquipment,
    property_plant_equipment_prompt,
)
from models.short_term_notes import (
    ShortTermNotesPayable,
    short_term_notes_payable_prompt,
)
from utils import get_company_info
from gemini_client import get_gemini

client = get_gemini()

# 添加 PDF_DIR 和 model_prompt_mapping 以便 GUI 使用
PDF_DIR = Path(__file__).parent.parent / "assets/pdfs"
RESULTS_DIR = Path(__file__).parent.parent / "assets/results"
REPORTS_DIR = Path(__file__).parent.parent / "assets/reports"
TEMPLATE_PATH = Path(__file__).parent.parent / "assets/template.xlsx"

model_prompt_mapping = {
    "cash_equivalents": {
        "prompt": cash_equivalents_prompt,
        "model": CashAndEquivalents,
    },
    # "total_liabilities": {
    #     "prompt": total_liabilities_prompt,
    #     "model": TotalLiabilities,
    # },
    # "prepayments": {
    #     "prompt": prepayments_prompt,
    #     "model": PrePayments,
    # },
    # "receivables_related_parties": {
    #     "prompt": receivables_related_parties_prompt,
    #     "model": ReceivablesRelatedParties,
    # },
    # "corporate_bond_payable": {
    #     "prompt": corporate_bond_payable_prompt,
    #     "model": CorporateBondPayable,
    # },
    # "property_plant_equipment": {
    #     "prompt": property_plant_equipment_prompt,
    #     "model": PropertyPlantEquipment,
    # },
    # "short_term_notes_payable": {
    #     "prompt": short_term_notes_payable_prompt,
    #     "model": ShortTermNotesPayable,
    # },
}


# 財務報表位置模型
class FinancialStatementLocation(BaseModel):
    """財務報表項目位置"""

    item_name: str = Field(description="財務報表項目名稱")
    page_numbers: list[int] = Field(description="該項目所在的頁數列表")
    found: bool = Field(description="是否找到該項目")


class FinancialStatementsAnalysis(BaseModel):
    """財務報表分析結果"""

    company_name: str = Field(description="公司名稱")
    pdf_year: int = Field(description="年份")
    individual_balance_sheet: FinancialStatementLocation = Field(
        description="個體資產負債表"
    )
    individual_comprehensive_income: FinancialStatementLocation = Field(
        description="個體綜合損益表"
    )
    individual_equity_changes: FinancialStatementLocation = Field(
        description="個體權益變動表"
    )
    individual_cash_flow: FinancialStatementLocation = Field(
        description="個體現金流量表"
    )
    important_accounting_items: FinancialStatementLocation = Field(
        description="重要會計項目明細表"
    )

    def get_all_page_numbers(self) -> list[int]:
        """
        提取所有財務報表的頁碼

        回傳：
            list[int]: 排序後的不重複頁碼列表
        """
        all_pages = []
        statements = [
            self.individual_balance_sheet,
            self.individual_comprehensive_income,
            self.individual_equity_changes,
            self.individual_cash_flow,
            self.important_accounting_items,
        ]

        for statement in statements:
            if statement.found and statement.page_numbers:
                all_pages.extend(statement.page_numbers)

        return sorted(list(set(all_pages)))


# 檢查是否是掃描檔
def check_scanned_pages(
    pdf_path: str, pages_to_check: list[int] | None = None, text_threshold: int = 1
) -> dict[int, bool]:
    """
    檢查指定的頁碼是否為「掃描影像頁」(沒有可複製文字)。

    參數：
      - pdf_path (str)：PDF 檔案路徑。
      - pages_to_check (list[int])：要檢查的頁面列表（以 1 開始的頁碼）。
      - text_threshold (int)：判斷文字頁的門檻（字元數），預設 1。如果回傳的文字長度 >= text_threshold 就判斷為有文字層。

    回傳：
      - dict[int, bool]：以頁碼為 key，value 為布林值。如果 True 則代表該頁「無可複製文字」，極可能是掃描影像頁；False 代表該頁有文字層，可複製文字。
    """
    results: dict[int, bool] = {}
    doc = fitz.open(pdf_path)
    total_pages = doc.page_count

    if pages_to_check is None:
        pages_to_check = range(1, total_pages + 1)

    for pg in pages_to_check:
        # 檢查頁碼是否合理
        if not (1 <= pg <= total_pages):
            raise ValueError(f"頁碼 {pg} 超出範圍（1~{total_pages}）。")

        page = doc.load_page(pg - 1)  # PyMuPDF 的頁面索引從 0 開始
        text = page.get_text("text") or ""  # 取出純文字層；如果 None，就改成空字串

        # 根據 text_threshold 判斷是否「無可複製文字」
        if len(text.strip()) < text_threshold:
            # 當文字長度小於門檻，就視為「掃描影像頁」
            results[pg] = True
        else:
            # 有足夠文字，就視為「文字頁」
            results[pg] = False

    doc.close()
    return results


def analyze_toc_and_extract_financial_statements(
    pdf_path: str,
) -> FinancialStatementsAnalysis:
    """
    分析PDF前5頁的目錄內容，找出各財務報表項目的頁數位置

    參數：
        pdf_path (str): PDF檔案路徑

    回傳：
        FinancialStatementsAnalysis: 包含各財務報表位置的分析結果
    """
    try:
        # 開啟PDF文件
        doc = fitz.open(pdf_path)
        total_pages = doc.page_count

        print(f"正在分析PDF文件：{pdf_path}")
        print(f"總頁數：{total_pages}")

        # 提取前5頁（或全部頁面，如果少於5頁）
        pages_to_extract = min(5, total_pages)
        print(f"提取前 {pages_to_extract} 頁進行目錄分析...")

        # 創建包含前5頁的新PDF
        pdf_writer = fitz.open()  # 創建新的PDF文檔

        for page_num in range(pages_to_extract):
            page = doc.load_page(page_num)
            pdf_writer.insert_pdf(doc, from_page=page_num, to_page=page_num)

        # 將前5頁轉換為base64
        pdf_bytes = pdf_writer.tobytes()
        base64_content = base64.b64encode(pdf_bytes).decode("utf-8")

        # 關閉文檔
        doc.close()
        pdf_writer.close()

        # 準備分析目錄的提示詞
        prompt = """
        請分析目錄頁，找出以下財務報表項目在目錄中顯示的頁數，並回傳公司名稱和年份：

        1. 個體資產負債表
        2. 個體綜合損益表  
        3. 個體權益變動表
        4. 個體現金流量表
        5. 重要會計項目明細表

        請仔細查看目錄中的項目名稱，可能會有類似的表達方式，例如：
        - "資產負債表"、"Balance Sheet"
        - "綜合損益表"、"Comprehensive Income Statement"
        - "權益變動表"、"Statement of Changes in Equity"
        - "現金流量表"、"Cash Flow Statement"
        - "重要會計項目明細表"、"Notes to Financial Statements"、"附註"

        注意：
        - 請根據目錄中顯示的頁數填寫
        - 如果找不到某個項目，請將found設為false
        - 如果某個報表跨越多頁，請列出所有相關頁數
        - 重要會計項目明細表不等於重要會計項目之說明，請注意區分
        - 請回傳公司名稱和年份
        """

        print("正在分析目錄頁內容...")

        result = client.call(prompt, FinancialStatementsAnalysis, base64_content)
        print("目錄分析完成！")

        return result

    except Exception as e:
        print(f"分析過程中發生錯誤：{str(e)}")
        raise e


def convert_pdf_to_markdown(
    pdf_path: str, pages: list[int], max_workers: int = 4
) -> dict[int, str]:
    """
    將指定頁數的PDF轉換為Markdown格式（使用多線程並行處理）

    參數：
        pdf_path (str): PDF檔案路徑
        pages (list[int]): 要轉換的頁數列表（1-based頁碼）
        max_workers (int): 最大線程數，預設為4

    回傳：
        dict[int, str]: 以頁碼為key，value為轉換後的Markdown內容
    """
    try:
        # 開啟PDF文件
        doc = fitz.open(pdf_path)
        total_pages = doc.page_count

        print(f"正在轉換PDF頁面 {pages} 為Markdown格式")
        valid_pages = [p for p in pages if 1 <= p <= total_pages]
        if not valid_pages:
            doc.close()
            raise ValueError(f"沒有有效的頁碼。有效範圍：1-{total_pages}")

        markdown_content = {}
        lock = threading.Lock()

        def process_single_page(page: int) -> tuple[int, str]:
            """處理單個頁面的轉換"""
            try:
                # 為每個頁面創建獨立的PDF文檔
                with lock:  # 保護PDF文檔操作
                    single_page_doc = fitz.open()
                    single_page_doc.insert_pdf(
                        doc, from_page=page - 1, to_page=page - 1
                    )
                    pdf_bytes = single_page_doc.tobytes()
                    single_page_doc.close()

                # 將單頁轉換為base64
                base64_content = base64.b64encode(pdf_bytes).decode("utf-8")

                # 準備轉換為Markdown的提示詞
                prompt = """
                請將這個PDF文件的內容轉換為Markdown格式。

                轉換要求：
                1. 保持表格結構，使用Markdown表格語法
                2. 保持標題層級結構
                3. 保持數字和文字的精確性
                4. 如果有財務數據表格，請確保數字對齊和格式正確
                5. 保持原始的段落分隔

                請直接返回Markdown格式的文本內容，不需要額外的格式包裝。
                """

                print(f"正在呼叫Gemini API轉換第 {page} 頁...")

                result = client.call(prompt, None, base64_content)

                print(f"第 {page} 頁轉換完成")
                return page, result

            except Exception as page_error:
                print(f"轉換第 {page} 頁時發生錯誤：{page_error}")
                error_content = f"# 第 {page} 頁轉換失敗\n\n轉換錯誤：{str(page_error)}"
                return page, error_content

        # 使用ThreadPoolExecutor進行並行處理
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有任務
            future_to_page = {
                executor.submit(process_single_page, page): page for page in valid_pages
            }

            # 收集結果
            for future in as_completed(future_to_page):
                page_num, content = future.result()
                markdown_content[page_num] = content

        # 關閉原始文檔
        doc.close()

        print(f"PDF轉Markdown完成！成功轉換了 {len(markdown_content)} 頁")
        return markdown_content

    except Exception as e:
        print(f"轉換PDF為Markdown時發生錯誤：{str(e)}")
        # 確保文檔被關閉
        try:
            if "doc" in locals():
                doc.close()
        except:
            pass
        raise e


@contextmanager
def temporary_files(*suffixes):
    """
    上下文管理器：創建多個臨時檔案，自動清理

    參數：
        *suffixes: 檔案後綴名列表

    產出：
        list[str]: 臨時檔案路徑列表
    """
    temp_files = []
    try:
        for suffix in suffixes:
            temp_file = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
            temp_files.append(temp_file.name)
            temp_file.close()
        yield temp_files
    finally:
        # 清理所有臨時檔案
        for temp_path in temp_files:
            try:
                os.unlink(temp_path)
            except (OSError, FileNotFoundError):
                pass  # 檔案可能已被刪除或不存在


def convert_markdown_to_pdf(markdown_content: dict[int, str], pdf_path: str) -> str:
    """
    將Markdown內容轉換為PDF頁面，並替換原始PDF中的對應頁面

    參數：
        markdown_content (dict[int, str]): 以頁碼為key，value為Markdown內容
        pdf_path (str): 原始PDF檔案路徑

    回傳：
        str: 新PDF檔案的路徑
    """
    try:
        if not markdown_content:
            print("沒有Markdown內容需要轉換")
            return pdf_path

        print(f"正在使用 Spire.Doc 將Markdown內容轉換為PDF並替換掃描頁面...")

        # 開啟原始PDF
        original_doc = fitz.open(pdf_path)
        total_pages = original_doc.page_count

        # 創建新的PDF文檔
        new_doc = fitz.open()
        from spire.doc import Document, FileFormat

        try:
            # 逐頁處理
            for page_num in range(1, total_pages + 1):
                if page_num in markdown_content:
                    print(f"正在轉換第 {page_num} 頁的Markdown內容...")

                    # 使用上下文管理器處理臨時檔案
                    with temporary_files(".md", ".pdf") as (
                        temp_md_path,
                        temp_pdf_path,
                    ):
                        # 寫入Markdown內容
                        with open(temp_md_path, "w", encoding="utf-8") as f:
                            f.write(markdown_content[page_num])

                        # 創建Word文檔並轉換
                        word_doc = Document()
                        try:
                            word_doc.LoadFromFile(temp_md_path)
                            word_doc.SaveToFile(temp_pdf_path, FileFormat.PDF)
                        finally:
                            word_doc.Dispose()

                        # 讀取生成的PDF並插入到新文檔中
                        converted_pdf = fitz.open(temp_pdf_path)
                        try:
                            new_doc.insert_pdf(converted_pdf)
                        finally:
                            converted_pdf.close()

                    print(f"第 {page_num} 頁轉換完成")

                else:
                    # 保留原始頁面
                    new_doc.insert_pdf(
                        original_doc, from_page=page_num - 1, to_page=page_num - 1
                    )

            # 生成新的檔案名稱
            original_path = Path(pdf_path)
            new_pdf_path = (
                original_path.parent
                / f"{original_path.stem}_converted{original_path.suffix}"
            )

            # 保存新的PDF
            new_doc.save(str(new_pdf_path))
            return str(new_pdf_path)

        finally:
            # 確保文檔被關閉
            new_doc.close()
            original_doc.close()

    except Exception as e:
        print(f"轉換Markdown為PDF時發生錯誤：{str(e)}")
        import sys

        sys.exit(1)


def genetate_verification_report(
    results: dict[str, BaseModel], pdf_data, filepath: Path
) -> str:
    # 將results轉換為易讀的文字格式
    results_text = "\n\n## 已提取的財務數據：\n"
    for model_name, data in results.items():
        results_text += f"\n### {model_name}:\n"
        if data:
            results_text += data.model_dump_json()
        else:
            results_text += "提取失敗或無數據"
        results_text += "\n"

    # 檢查和驗證results
    prompt = f"""
    你是專業的財務分析師，請分析這個財務報表PDF文件，並根據已提取的數據進行驗證和補充分析。

    ## 任務要求：
    1. **驗證已提取數據的準確性**：檢查下方提取的財務數據是否與PDF中的原始數據一致
    2. **找出遺漏或錯誤的項目**：識別可能被遺漏或提取錯誤的財務數據
    3. **數據一致性檢查**：確認各財務報表項目之間的邏輯一致性

    ## 分析重點：
    - 項目的完整性和準確性
    - 分類和金額正確性  
    - reason 詳細分析

    ## 參考標準：
    - 確保數字精確到小數點
    - 注意貨幣單位（千元、萬元等）
    - 檢查期間比較數據的一致性
    - 驗證會計科目分類的正確性
    - 優先以大表的數據為主，附註的數據為輔

    {results_text}

    ## 輸出格式：
    請以Markdown的方式使用繁體中文回應，包含：
    1. **數據驗證結果**：已提取數據的準確性評估
    2. **發現的問題**：錯誤、遺漏或不一致之處
    3. **補充資訊**：PDF中其他重要的財務資訊
    4. **建議改進**：數據提取可以改進的地方
    """

    # 保存驗證結果
    result = client.call(prompt, None, pdf_data)

    # 確保報告目錄存在
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # 使用 REPORTS_DIR 保存驗證報告
    report_path = REPORTS_DIR / f"{filepath.stem}_verification.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"# {filepath.stem} 財務數據驗證報告\n\n")
        f.write(result)

    print(f"✓ 財務數據驗證完成，報告已保存至：{report_path}")
    return str(report_path)


def process_single_pdf_with_gemini(
    filepath: Path, model_selection: list[str]
) -> tuple[dict, str]:
    """
    使用 Gemini 處理單個 PDF 檔案的財務報表分析

    參數:
        filepath: PDF 檔案路徑
        model_selection: 選擇的模型列表

    回傳:
        tuple[dict, str]: 包含所有模型分析結果的字典、驗證報告的路徑
    """
    try:
        # 分析目錄並提取財務報表位置
        toc_content = analyze_toc_and_extract_financial_statements(filepath)

        # 正確地提取所有財務報表的頁碼
        table_pages = toc_content.get_all_page_numbers()
        # 檢查是否為掃描文件
        scan_results = check_scanned_pages(filepath, table_pages) if table_pages else {}
        scanned_pages = [page for page, is_scan in scan_results.items() if is_scan]

        # 處理掃描頁面
        if scanned_pages:
            try:
                markdown_content = convert_pdf_to_markdown(str(filepath), scanned_pages)
                new_pdf_path = convert_markdown_to_pdf(markdown_content, str(filepath))
                filepath = Path(new_pdf_path)
                print(f"已處理掃描頁面並生成新PDF: {new_pdf_path}")
            except Exception as e:
                print(f"轉換掃描頁面失敗：{e}")
                # 繼續使用原始PDF

        # 讀取PDF並轉換為base64
        pdf_data = base64.b64encode(filepath.read_bytes()).decode("utf-8")
        pdf_info = get_company_info(toc_content.company_name, toc_content.pdf_year)

        def process_model(prompt_model_pair) -> tuple[str, BaseModel | None]:
            """處理單個模型的分析"""
            model, prompt = prompt_model_pair
            prompt = prompt + f"\n\n資料請優先以大表為主，附註為輔"
            try:
                result: BaseModel = client.call(prompt, model, pdf_data)
                # 直接返回模型對象而不是字典
                return model.__name__, result
            except Exception as e:
                print(f"{model.__name__} 分析失敗：{e}")
                return model.__name__, None

        # model_prompt_mapping
        model_pairs = [
            (
                model_prompt_mapping[model_name]["model"],
                model_prompt_mapping[model_name]["prompt"],
            )
            for model_name in model_prompt_mapping.keys()
            if not model_selection or model_name in model_selection
        ]

        results: dict[str, BaseModel] = {}
        results["pdf_info"] = pdf_info
        # 使用ThreadPoolExecutor進行並行處理
        with ThreadPoolExecutor(max_workers=4) as executor:
            # 提交所有任務
            future_to_model = {
                executor.submit(process_model, pair): pair[0].__name__
                for pair in model_pairs
            }

            # 收集結果
            for future in as_completed(future_to_model):
                model_name = future_to_model[future]
                try:
                    result_name, result_data = future.result()
                    if result_data is not None:
                        results[result_name] = result_data
                        print(f"✓ {model_name} 處理完成")
                    else:
                        print(f"✗ {model_name} 處理失敗")
                except Exception as e:
                    print(f"✗ {model_name} 處理異常：{e}")

        verification_report_path = genetate_verification_report(
            results, pdf_data, filepath
        )

        return results, verification_report_path

    except Exception as e:
        print(f"處理 PDF 檔案時發生錯誤：{e}")
        return {}, ""


def export_excel(results: dict[str, BaseModel], filepath: Path):
    # 檢查模板檔案的格式
    # if TEMPLATE_PATH.suffix.lower() == ".xls":
    #     # 處理 .xls 格式
    #     import pandas as pd
    #     from openpyxl import Workbook

    #     # 使用 pandas 讀取 .xls 檔案的所有工作表
    #     df_dict = pd.read_excel(TEMPLATE_PATH, sheet_name=None)

    #     # 創建新的 openpyxl 工作簿
    #     wb = Workbook()
    #     # 移除默認的工作表
    #     wb.remove(wb.active)

    #     # 將每個工作表轉換為 openpyxl 格式
    #     for sheet_name, df in df_dict.items():
    #         ws = wb.create_sheet(title=sheet_name)
    #         # 將 DataFrame 寫入工作表
    #         for row_idx, row in enumerate(df.values, 1):
    #             for col_idx, value in enumerate(row, 1):
    #                 ws.cell(
    #                     row=row_idx + 1, column=col_idx, value=value
    #                 )  # +1 因為要保留標題行
    #         # 寫入標題行
    #         for col_idx, col_name in enumerate(df.columns, 1):
    #             ws.cell(row=1, column=col_idx, value=col_name)
    # else:
    #     # 處理 .xlsx 格式（原有邏輯）
    #     wb = load_workbook(TEMPLATE_PATH, keep_vba=True)

    wb = load_workbook(TEMPLATE_PATH)

    for model in results.values():
        # 把同一個 worksheet 傳給每個 model 寫入
        model.fill_excel(wb)

    # 全部填完後再存檔
    wb.save("output.xlsx")


# 如果直接執行此檔案，運行測試
if __name__ == "__main__":
    # 重置token計數器
    # token_tracker.reset()

    # 測試用的 PDF 檔案
    test_files = [
        # "quartely-results-2024-zh_tcm27-94407.pdf",
        "fin_202503071324328842.pdf"
    ]

    for filename in test_files:
        filepath = PDF_DIR / filename
        if filepath.exists():
            print(f"測試處理檔案: {filename}")
            results, verification_report_path = process_single_pdf_with_gemini(
                filepath, model_selection=model_prompt_mapping.keys()
            )
            export_excel(results, "")
            # 確保結果目錄存在
            RESULTS_DIR.mkdir(parents=True, exist_ok=True)

            # 使用 RESULTS_DIR 保存結果
            results_path = RESULTS_DIR / f"{filepath.stem}_gemini_results.json"
            with open(results_path, "w", encoding="utf-8") as f:
                json_results = {}
                for model_name, result in results.items():
                    if hasattr(result, "model_dump"):
                        json_results[model_name] = result.model_dump()
                    else:
                        json_results[model_name] = str(result)
                json.dump(json_results, f, indent=4, ensure_ascii=False)

            print(f"✓ 結果已保存到: {results_path}")
